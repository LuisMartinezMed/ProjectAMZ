from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable, Optional

import openpyxl
from openpyxl.utils.cell import range_boundaries
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..models import Listing
from ..utils.dates import parse_date
from ..utils.money import parse_percentage, q4, to_decimal
from ..utils.validators import ValidationError
from .costs_service import CostsService
from .inventory_service import InventoryService
from .products_service import ProductsService, infer_units_per_listing
from .sales_service import SalesService


@dataclass(frozen=True)
class WorkbookPreview:
    path: str
    sheets: list[dict[str, Any]]
    catalog_rows: int
    sales_rows: int
    analysis_rows: int
    simulator_rows: int
    warnings: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class ImportResult:
    products_created: int = 0
    listings_created: int = 0
    cost_versions_created: int = 0
    fee_versions_created: int = 0
    price_versions_created: int = 0
    inventory_batches_created: int = 0
    sales_imported: int = 0
    rows_skipped: int = 0
    warnings: list[str] = field(default_factory=list)


class ImportExportService:
    def __init__(self, session: Session):
        self.session = session
        self.products = ProductsService(session)
        self.costs = CostsService(session)
        self.inventory = InventoryService(session)
        self.sales = SalesService(session)

    def preview_workbook(self, workbook_path: str | Path) -> WorkbookPreview:
        workbook_path = Path(workbook_path)
        wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=False)
        sheets = []
        for ws in wb.worksheets:
            tables = []
            for table_name in ws.tables.keys():
                table = ws.tables[table_name]
                min_col, min_row, max_col, _ = range_boundaries(table.ref)
                headers = [ws.cell(min_row, col).value for col in range(min_col, max_col + 1)]
                tables.append({"name": table_name, "ref": table.ref, "headers": headers})
            sheets.append({"name": ws.title, "dimension": ws.calculate_dimension(), "tables": tables})

        catalog_rows = len(list(self._table_rows(wb, "Catalogo y Costos", "Costos_Catalogo")))
        sales = list(self._table_rows(wb, "Registro de Ventas", "Ventas"))
        analysis_rows = len(list(self._table_rows(wb, "Analisis de productos", "TB_Analisis")))
        simulator_rows = len(list(self._table_rows(wb, "Simulador de compraventa", "Tabla_Simulador")))
        warnings = []
        missing_net = sum(1 for row in sales if self._row_value(row, "Ingreso Neto") in (None, ""))
        missing_sku = sum(1 for row in sales if self._row_value(row, "SKU") in (None, ""))
        if missing_net:
            warnings.append(f"{missing_net} sales rows are missing net received and need review or fee estimates.")
        if missing_sku:
            warnings.append(f"{missing_sku} sales rows are missing SKU.")
        if catalog_rows or analysis_rows:
            warnings.append(
                "Costo Total ($) and Costo total por paquete (Luna+AMZ) include Amazon fees. "
                "They are not used as FIFO inventory COGS when actual net received is available."
            )
        return WorkbookPreview(
            path=str(workbook_path),
            sheets=sheets,
            catalog_rows=catalog_rows,
            sales_rows=len(sales),
            analysis_rows=analysis_rows,
            simulator_rows=simulator_rows,
            warnings=warnings,
        )

    def import_workbook(
        self,
        workbook_path: str | Path,
        *,
        import_catalog: bool = True,
        import_initial_stock: bool = True,
        import_sales: bool = False,
        missing_net_strategy: str = "skip",
        effective_from: Optional[date] = None,
    ) -> ImportResult:
        workbook_path = Path(workbook_path)
        wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=False)
        effective_from = effective_from or date(1900, 1, 1)
        counts = {
            "products_created": 0,
            "listings_created": 0,
            "cost_versions_created": 0,
            "fee_versions_created": 0,
            "price_versions_created": 0,
            "inventory_batches_created": 0,
            "sales_imported": 0,
            "rows_skipped": 0,
        }
        warnings: list[str] = []
        marketplace = self.products.get_marketplace("Amazon")
        analysis_lookup = self._build_analysis_lookup(wb)

        catalog_rows = self._table_rows(wb, "Catalogo y Costos", "Costos_Catalogo") if import_catalog else []
        for row in catalog_rows:
            sku = self._clean_text(self._row_value(row, "SKU"))
            asin = self._clean_text(self._row_value(row, "ASIN"))
            name = self._clean_text(self._row_value(row, "Nombre del Producto"))
            if not sku or not name:
                counts["rows_skipped"] += 1
                warnings.append(f"Skipped catalog row with missing SKU/name: {row}")
                continue

            analysis = self._analysis_for(analysis_lookup, sku=sku, asin=asin)
            analysis_units = analysis.get("units_per_listing") if analysis else None
            units_per_listing = int(analysis_units or infer_units_per_listing(name))

            existing = self.session.scalar(
                select(Listing).where(Listing.marketplace_id == marketplace.id, Listing.sku == sku)
            )
            if existing:
                listing = existing
                listing.asin = asin or listing.asin
                listing.listing_name = name
                listing.units_per_listing = units_per_listing
            else:
                base_product = self.products.create_base_product(name=name)
                listing = self.products.create_listing(
                    marketplace_id=marketplace.id,
                    base_product_id=base_product.id,
                    sku=sku,
                    asin=asin,
                    listing_name=name,
                    units_per_listing=units_per_listing,
                )
                counts["products_created"] += 1
                counts["listings_created"] += 1

            (
                supplier_cost_per_listing,
                supplier_unit_cost,
                warehouse_cost_per_unit,
                total_landed_cost_per_unit,
                cost_notes,
            ) = self._internal_landed_costs(
                analysis=analysis,
                sku=sku,
                units_per_listing=units_per_listing,
                warnings=warnings,
            )
            self.costs.create_product_cost_version(
                base_product_id=listing.base_product_id,
                effective_from=effective_from,
                supplier_lot_cost=supplier_cost_per_listing,
                supplier_lot_units=units_per_listing,
                supplier_unit_cost=supplier_unit_cost,
                warehouse_cost_per_unit=warehouse_cost_per_unit,
                total_landed_cost_per_unit=total_landed_cost_per_unit,
                notes=cost_notes,
            )
            counts["cost_versions_created"] += 1

            price = self._row_value(row, "Precio Venta ($)")
            if price not in (None, ""):
                self.costs.create_price_version(
                    listing_id=listing.id,
                    marketplace_id=marketplace.id,
                    effective_from=effective_from,
                    suggested_selling_price=price,
                    buybox_percentage=parse_percentage(self._row_value(row, "BuyBox Mensual (%)") or 0),
                    notes="Imported from Catalogo y Costos",
                )
                counts["price_versions_created"] += 1

            stock_listings = int(self._row_value(row, "Stock FBA Inicial") or 0)
            if import_initial_stock and stock_listings > 0:
                self.inventory.create_initial_stock(
                    base_product_id=listing.base_product_id,
                    received_date=effective_from,
                    physical_units=stock_listings * units_per_listing,
                    unit_landed_cost=total_landed_cost_per_unit,
                    notes="Imported initial FBA stock with internal landed cost",
                )
                counts["inventory_batches_created"] += 1

        if import_catalog:
            counts["fee_versions_created"] += self._import_fee_versions_from_analysis(
                wb, marketplace.id, effective_from
            )

        if import_sales:
            sales_result = self._import_sales(wb, missing_net_strategy=missing_net_strategy)
            counts["sales_imported"] += sales_result[0]
            counts["rows_skipped"] += sales_result[1]
            warnings.extend(sales_result[2])

        return ImportResult(**counts, warnings=warnings)

    def _import_fee_versions_from_analysis(self, wb, marketplace_id: int, effective_from: date) -> int:
        created = 0
        for row in self._table_rows(wb, "Analisis de productos", "TB_Analisis"):
            asin = self._clean_text(self._row_value(row, "ASIN"))
            if not asin:
                continue
            listing = self.session.scalar(
                select(Listing).where(Listing.marketplace_id == marketplace_id, Listing.asin == asin)
            )
            if listing is None:
                continue
            fees = q4(self._row_value(row, "Fees de Amazon") or 0)
            shipping = q4(self._row_value(row, "Envio de Amazon") or 0)
            self.costs.create_marketplace_fee_version(
                listing_id=listing.id,
                marketplace_id=marketplace_id,
                effective_from=effective_from,
                referral_fee=fees,
                amazon_shipping_or_fulfillment_fee=shipping,
                other_marketplace_fee=0,
                notes="Imported from Analisis de productos",
            )
            created += 1
        return created

    def _import_sales(self, wb, *, missing_net_strategy: str) -> tuple[int, int, list[str]]:
        imported = 0
        skipped = 0
        warnings: list[str] = []
        if missing_net_strategy not in {"skip", "estimate"}:
            raise ValidationError("missing_net_strategy must be 'skip' or 'estimate'")
        for row_number, row in enumerate(self._table_rows(wb, "Registro de Ventas", "Ventas"), start=2):
            sale_date = parse_date(self._row_value(row, "Fecha"))
            sku = self._clean_text(self._row_value(row, "SKU"))
            quantity = self._row_value(row, "Unidades Vendidas")
            gross = self._row_value(row, "Ingreso Bruto ($)")
            net = self._row_value(row, "Ingreso Neto")
            if not sale_date or not sku or not quantity or gross in (None, ""):
                skipped += 1
                warnings.append(f"Skipped sales row {row_number}: missing required sale date/SKU/quantity/gross.")
                continue
            if net in (None, "") and missing_net_strategy == "skip":
                skipped += 1
                warnings.append(f"Skipped sales row {row_number}: missing net received.")
                continue
            try:
                self.sales.add_manual_sale(
                    sale_date=sale_date,
                    sku=sku,
                    quantity_listings_sold=int(quantity),
                    gross_revenue=to_decimal(gross),
                    net_received=None if net in (None, "") else to_decimal(net),
                    source="excel_import",
                    notes="Imported from Registro de Ventas",
                )
                imported += 1
            except Exception as exc:
                skipped += 1
                warnings.append(f"Skipped sales row {row_number}: {exc}")
        return imported, skipped, warnings

    def _table_rows(self, wb, sheet_name: str, table_name: str) -> Iterable[dict[str, Any]]:
        ws = self._worksheet_for_table(wb, sheet_name, table_name)
        if ws is None or table_name not in ws.tables:
            return []
        table = ws.tables[table_name]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [ws.cell(min_row, col).value for col in range(min_col, max_col + 1)]
        rows = []
        for row_idx in range(min_row + 1, max_row + 1):
            values = [ws.cell(row_idx, col).value for col in range(min_col, max_col + 1)]
            if not any(value not in (None, "") for value in values):
                continue
            rows.append(dict(zip(headers, values)))
        return rows

    def _worksheet_for_table(self, wb, sheet_name: str, table_name: str):
        if sheet_name in wb.sheetnames and table_name in wb[sheet_name].tables:
            return wb[sheet_name]
        target_sheet = self._normalize_label(sheet_name)
        for ws in wb.worksheets:
            if table_name in ws.tables:
                return ws
            if self._normalize_label(ws.title) == target_sheet:
                return ws
        return None

    def _build_analysis_lookup(self, wb) -> dict[str, dict[str, Any]]:
        lookup: dict[str, dict[str, Any]] = {}
        for row in self._table_rows(wb, "Analisis de productos", "TB_Analisis"):
            sku = self._clean_text(self._row_value(row, "SKU"))
            asin = self._clean_text(self._row_value(row, "ASIN"))
            units = self._row_value(row, "Unidades por paquete Amazon")
            info = {
                "sku": sku,
                "asin": asin,
                "units_per_listing": int(units) if units not in (None, "") else None,
                "supplier_unit_cost": self._row_value(row, "Precio por unidad"),
                "supplier_cost_per_listing": self._row_value(row, "Precio por paquete"),
                "warehouse_cost_per_listing": self._row_value(row, "Costo almacen Luna"),
                "internal_landed_cost_per_listing": self._row_value(row, "Gastos totales"),
                "amazon_total_cost_per_listing": self._row_value(row, "Costo total por paquete (Luna+AMZ)"),
                "referral_fee": self._row_value(row, "Fees de Amazon"),
                "amazon_shipping_or_fulfillment_fee": self._row_value(row, "Envio de Amazon"),
            }
            if sku:
                lookup[f"sku:{sku.upper()}"] = info
            if asin:
                lookup[f"asin:{asin.upper()}"] = info
        return lookup

    def _analysis_for(
        self, lookup: dict[str, dict[str, Any]], *, sku: Optional[str], asin: Optional[str]
    ) -> Optional[dict[str, Any]]:
        if sku:
            match = lookup.get(f"sku:{sku.upper()}")
            if match:
                return match
        if asin:
            return lookup.get(f"asin:{asin.upper()}")
        return None

    def _internal_landed_costs(
        self,
        *,
        analysis: Optional[dict[str, Any]],
        sku: str,
        units_per_listing: int,
        warnings: list[str],
    ) -> tuple[Decimal, Decimal, Decimal, Decimal, str]:
        supplier_unit_cost = self._decimal_or_none(analysis.get("supplier_unit_cost")) if analysis else None
        supplier_cost_per_listing = (
            self._decimal_or_none(analysis.get("supplier_cost_per_listing")) if analysis else None
        )
        warehouse_cost_per_listing = (
            self._decimal_or_none(analysis.get("warehouse_cost_per_listing")) if analysis else None
        )
        internal_landed_cost_per_listing = (
            self._decimal_or_none(analysis.get("internal_landed_cost_per_listing")) if analysis else None
        )

        if supplier_unit_cost is None and supplier_cost_per_listing is not None:
            supplier_unit_cost = q4(supplier_cost_per_listing / units_per_listing)
        if supplier_cost_per_listing is None and supplier_unit_cost is not None:
            supplier_cost_per_listing = q4(supplier_unit_cost * units_per_listing)

        warehouse_cost_per_listing = warehouse_cost_per_listing or Decimal("0")
        if internal_landed_cost_per_listing is None and supplier_cost_per_listing is not None:
            internal_landed_cost_per_listing = q4(supplier_cost_per_listing + warehouse_cost_per_listing)

        notes = "Imported internal landed cost from Analisis de productos"
        if internal_landed_cost_per_listing is None:
            internal_landed_cost_per_listing = Decimal("0")
            notes = "Imported with missing internal landed cost; review before sales import"
            warnings.append(
                f"No Analisis de productos internal landed cost was found for SKU {sku}; "
                "FIFO inventory cost was set to 0 instead of using Amazon-inclusive Costo Total ($)."
            )

        if supplier_unit_cost is None:
            supplier_unit_cost = q4(internal_landed_cost_per_listing / units_per_listing)
        if supplier_cost_per_listing is None:
            supplier_cost_per_listing = q4(supplier_unit_cost * units_per_listing)

        warehouse_cost_per_unit = q4(warehouse_cost_per_listing / units_per_listing)
        total_landed_cost_per_unit = q4(internal_landed_cost_per_listing / units_per_listing)
        return (
            q4(supplier_cost_per_listing),
            q4(supplier_unit_cost),
            warehouse_cost_per_unit,
            total_landed_cost_per_unit,
            notes,
        )

    @staticmethod
    def _decimal_or_none(value: Any) -> Optional[Decimal]:
        if value in (None, ""):
            return None
        return q4(value)

    def _row_value(self, row: dict[str, Any], *candidates: str) -> Any:
        for candidate in candidates:
            if candidate in row:
                return row[candidate]
        normalized = {self._normalize_label(key): value for key, value in row.items()}
        for candidate in candidates:
            key = self._normalize_label(candidate)
            if key in normalized:
                return normalized[key]
        return None

    @staticmethod
    def _clean_text(value: Any) -> Optional[str]:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    @staticmethod
    def _normalize_label(value: Any) -> str:
        text = "" if value is None else str(value)
        text = unicodedata.normalize("NFKD", text)
        text = "".join(char for char in text if not unicodedata.combining(char))
        return re.sub(r"[^a-z0-9]+", "", text.lower())
