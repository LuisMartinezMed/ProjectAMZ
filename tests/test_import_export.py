from __future__ import annotations

from datetime import date
from decimal import Decimal

import openpyxl
from openpyxl.worksheet.table import Table
from sqlalchemy import select

from ecommerce_manager.app.models import (
    InventoryBatch,
    Listing,
    MarketplaceFeeVersion,
    ProductCostVersion,
    SaleLine,
)
from ecommerce_manager.app.services.import_export_service import ImportExportService
from ecommerce_manager.app.utils.money import parse_percentage


def test_parse_percentage_accepts_decimal_whole_and_percent_text():
    assert parse_percentage("0.21") == Decimal("0.2100")
    assert parse_percentage("21") == Decimal("0.2100")
    assert parse_percentage("21%") == Decimal("0.2100")


def test_catalog_import_uses_internal_landed_costs_for_fifo(tmp_path, session):
    workbook_path = tmp_path / "projectamz.xlsx"
    wb = openpyxl.Workbook()
    catalog = wb.active
    catalog.title = "Catalogo y Costos"
    catalog_headers = [
        "SKU",
        "ASIN",
        "Nombre del Producto",
        "Costo Total ($)",
        "Precio Venta ($)",
        "Stock FBA Inicial",
        "BuyBox Mensual (%)",
    ]
    catalog.append(catalog_headers)
    catalog_rows = [
        ["CB-FXA5-AL88", "B00FSYHBZ0", "Mitchum Power Gel", 11.2990, 17.67, 132, "21%"],
        ["A8-PPC9-SJRN", "B0071N7G2Y", "Febreze Car Vent Air Freshener, Midnight", 23.5184, 32.57, 90, "0"],
        ["7K-V0I3-W59G", "B07VRTH9PM", "Febreze Odor Clear Car Vent Air Freshener", 25.1840, 26.00, 90, "0"],
        ["IN-22OL-V8VV", "B0088PRSIQ", "Febreze Air Refresher", 19.6965, 26.00, 98, "0"],
        ["HG-B5SC-V1LB", "B07Y94Y3JL", "Mattel Games UNO Junior Card Game", 9.3350, 9.98, 120, "43"],
    ]
    for row in catalog_rows:
        catalog.append(row)
    catalog.add_table(Table(displayName="Costos_Catalogo", ref="A1:G6"))

    analysis = wb.create_sheet("Analisis de productos")
    analysis_headers = [
        "Nombre del producto",
        "Categoria del producto",
        "ASIN",
        "SKU",
        "Unidades por paquete Amazon",
        "Precio por unidad",
        "Precio por paquete",
        "Costo almacen Luna",
        "Gastos totales",
        "Costo total por paquete (Luna+AMZ)",
        "Fees de Amazon",
        "Envio de Amazon",
    ]
    analysis.append(analysis_headers)
    analysis_rows = [
        ["Mitchum", "Beauty", "B00FSYHBZ0", "CB-FXA5-AL88", 4, 1.95, 7.80, 1.00, 8.80, 11.2990, 2.499, 0],
        ["Febreze 8", "Automotive", "B0071N7G2Y", "A8-PPC9-SJRN", 8, 1.85, 14.80, 1.30, 16.10, 23.5184, 5.0, 2.4184],
        ["Febreze 8b", "Automotive", "B07VRTH9PM", "7K-V0I3-W59G", 8, 1.95, 15.60, 1.00, 16.60, 25.1840, 2.5, 0],
        ["Febreze 3", "Home", "B0088PRSIQ", "IN-22OL-V8VV", 3, 5.90, 17.70, 1.00, 18.70, 19.6965, 2.5, 0],
        ["UNO", "Toys", "B07Y94Y3JL", "HG-B5SC-V1LB", 1, 4.77, 4.77, 0.50, 5.27, 9.3350, 1.0, 0],
    ]
    for row in analysis_rows:
        analysis.append(row)
    analysis.add_table(Table(displayName="TB_Analisis", ref="A1:L6"))
    wb.save(workbook_path)

    result = ImportExportService(session).import_workbook(workbook_path, effective_from=date(2026, 1, 1))
    assert result.listings_created == 5

    expected_units = {
        "CB-FXA5-AL88": 4,
        "A8-PPC9-SJRN": 8,
        "7K-V0I3-W59G": 8,
        "IN-22OL-V8VV": 3,
        "HG-B5SC-V1LB": 1,
    }
    for sku, units in expected_units.items():
        listing = session.scalar(select(Listing).where(Listing.sku == sku))
        assert listing is not None
        assert listing.units_per_listing == units

    expected_costs = {
        "CB-FXA5-AL88": {
            "supplier_unit": Decimal("1.9500"),
            "supplier_lot": Decimal("7.8000"),
            "warehouse_unit": Decimal("0.2500"),
            "landed_unit": Decimal("2.2000"),
            "old_amazon_inclusive_unit": Decimal("2.8248"),
            "referral": Decimal("2.4990"),
            "shipping": Decimal("0.0000"),
        },
        "A8-PPC9-SJRN": {
            "supplier_unit": Decimal("1.8500"),
            "supplier_lot": Decimal("14.8000"),
            "warehouse_unit": Decimal("0.1625"),
            "landed_unit": Decimal("2.0125"),
            "old_amazon_inclusive_unit": Decimal("2.9398"),
            "referral": Decimal("5.0000"),
            "shipping": Decimal("2.4184"),
        },
        "HG-B5SC-V1LB": {
            "supplier_unit": Decimal("4.7700"),
            "supplier_lot": Decimal("4.7700"),
            "warehouse_unit": Decimal("0.5000"),
            "landed_unit": Decimal("5.2700"),
            "old_amazon_inclusive_unit": Decimal("9.3350"),
            "referral": Decimal("1.0000"),
            "shipping": Decimal("0.0000"),
        },
    }
    for sku, expected in expected_costs.items():
        listing = session.scalar(select(Listing).where(Listing.sku == sku))
        version = session.scalar(
            select(ProductCostVersion).where(ProductCostVersion.base_product_id == listing.base_product_id)
        )
        batch = session.scalar(
            select(InventoryBatch).where(InventoryBatch.base_product_id == listing.base_product_id)
        )
        fee = session.scalar(select(MarketplaceFeeVersion).where(MarketplaceFeeVersion.listing_id == listing.id))
        assert version.supplier_unit_cost == expected["supplier_unit"]
        assert version.supplier_lot_cost == expected["supplier_lot"]
        assert version.warehouse_cost_per_unit == expected["warehouse_unit"]
        assert version.total_landed_cost_per_unit == expected["landed_unit"]
        assert version.total_landed_cost_per_unit != expected["old_amazon_inclusive_unit"]
        assert batch.unit_landed_cost == expected["landed_unit"]
        assert batch.unit_landed_cost != expected["old_amazon_inclusive_unit"]
        assert fee.referral_fee == expected["referral"]
        assert fee.amazon_shipping_or_fulfillment_fee == expected["shipping"]
        assert fee.other_marketplace_fee == Decimal("0.0000")
        assert fee.estimated_total_fee == expected["referral"] + expected["shipping"]

    febreze = session.scalar(select(Listing).where(Listing.sku == "A8-PPC9-SJRN"))
    batch = session.scalar(
        select(InventoryBatch).where(InventoryBatch.base_product_id == febreze.base_product_id)
    )
    assert batch.initial_physical_units == 720
    assert batch.remaining_physical_units == 720
    assert batch.unit_landed_cost == Decimal("2.0125")


def test_sales_import_uses_ingreso_neto_without_double_subtracting_fees(tmp_path, session):
    workbook_path = tmp_path / "sales_net.xlsx"
    wb = openpyxl.Workbook()
    catalog = wb.active
    catalog.title = "Catalogo y Costos"
    catalog.append(
        [
            "SKU",
            "ASIN",
            "Nombre del Producto",
            "Costo Total ($)",
            "Precio Venta ($)",
            "Stock FBA Inicial",
            "BuyBox Mensual (%)",
        ]
    )
    catalog.append(
        ["A8-PPC9-SJRN", "B0071N7G2Y", "Febreze Car Vent Air Freshener, Midnight", 23.5184, 32.57, 90, "0"]
    )
    catalog.add_table(Table(displayName="Costos_Catalogo", ref="A1:G2"))

    analysis = wb.create_sheet("Analisis de productos")
    analysis.append(
        [
            "Nombre del producto",
            "Categoria del producto",
            "ASIN",
            "SKU",
            "Unidades por paquete Amazon",
            "Precio por unidad",
            "Precio por paquete",
            "Costo almacen Luna",
            "Gastos totales",
            "Costo total por paquete (Luna+AMZ)",
            "Fees de Amazon",
            "Envio de Amazon",
        ]
    )
    analysis.append(
        ["Febreze 8", "Automotive", "B0071N7G2Y", "A8-PPC9-SJRN", 8, 1.85, 14.80, 1.30, 16.10, 23.5184, 5.0, 2.4184]
    )
    analysis.add_table(Table(displayName="TB_Analisis", ref="A1:L2"))

    sales = wb.create_sheet("Registro de Ventas")
    sales.append(["Fecha", "SKU", "Producto", "Unidades Vendidas", "Ingreso Bruto ($)", "Ingreso Neto", "Mes", "Ano"])
    sales.append([date(2026, 2, 1), "A8-PPC9-SJRN", "Febreze", 1, 32.57, 20.00, "febrero", 2026])
    sales.add_table(Table(displayName="Ventas", ref="A1:H2"))
    wb.save(workbook_path)

    result = ImportExportService(session).import_workbook(
        workbook_path,
        import_sales=True,
        effective_from=date(2026, 1, 1),
    )

    line = session.scalar(select(SaleLine))
    assert result.sales_imported == 1
    assert line.net_received == Decimal("20.0000")
    assert line.net_received_estimated is False
    assert line.total_cogs_snapshot == Decimal("16.1000")
    assert line.marketplace_fee_snapshot == Decimal("7.4184")
    assert line.profit_snapshot == Decimal("3.9000")
